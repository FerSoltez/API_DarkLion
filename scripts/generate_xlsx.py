"""
generate_xlsx.py — Script para generar órdenes de producción (.xlsx)
Recibe datos por JSON y usa openpyxl para escribir en la plantilla.
Llena PORTADA (datos generales + matriz de tallas) y LISTA 1 (listado
de nombres/tallas/cantidades). Si hay más de 25 items, crea hojas
adicionales LISTA 2, LISTA 3, etc.

Uso: python generate_xlsx.py <input_json_path> <output_xlsx_path>
"""
import sys
import json
import os
import copy
import io
import tempfile
import urllib.request
import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
from openpyxl.utils.units import points_to_pixels, pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D

# Mapa de tallas: tipo -> talla -> celda de Excel
MAPA_TALLAS = {
    'DAMA': {
        'XCH': 'F22', 'CH': 'H22', 'MD': 'J22', 'GD': 'L22',
        'XL': 'N22', 'XXL': 'P22', 'XXXL': 'R22',
    },
    'CABALLERO': {
        'XCH': 'F23', 'CH': 'H23', 'MD': 'J23', 'GD': 'L23',
        'XL': 'N23', 'XXL': 'P23', 'XXXL': 'R23',
    },
    'INFANTIL': {
        'XCH': 'F21', 'CH': 'H21', 'MD': 'J21', 'GD': 'L21',
        'XL': 'N21', 'XXL': 'P21', 'XXXL': 'R21',
    },
}

# Constantes para la hoja LISTA
LISTA_FIRST_ROW = 9
LISTA_LAST_ROW = 33
LISTA_ITEMS_PER_SHEET = 25  # filas 9-33


def _column_width_to_pixels(width):
    """Convierte ancho de columna de Excel a pixeles (aproximado)."""
    if width is None:
        width = 8.43
    return int((float(width) + 0.75) * 7)


def _range_size_pixels(ws, start_col, end_col, start_row, end_row):
    """Calcula tamaño en pixeles de un rango rectangular."""
    total_width = 0
    for col_idx in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        col_dim = ws.column_dimensions.get(letter)
        total_width += _column_width_to_pixels(col_dim.width if col_dim else None)

    total_height = 0
    for row_idx in range(start_row, end_row + 1):
        row_dim = ws.row_dimensions.get(row_idx)
        row_height_points = row_dim.height if row_dim and row_dim.height is not None else 15
        total_height += int(points_to_pixels(row_height_points))

    return total_width, total_height


def _download_image_to_temp(image_url):
    """Descarga una imagen HTTP/HTTPS a un archivo temporal y devuelve su ruta."""
    suffix = '.png'
    lower_url = image_url.lower()
    if '.jpg' in lower_url or '.jpeg' in lower_url:
        suffix = '.jpg'
    elif '.webp' in lower_url:
        suffix = '.webp'

    with urllib.request.urlopen(image_url, timeout=20) as response:
        data = response.read()

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp_file.write(data)
    temp_file.flush()
    temp_file.close()
    return temp_file.name


def _insert_centered_image(ws, image_path):
    """Inserta imagen centrada dentro del bloque A29:L46 en la hoja PORTADA."""
    img = XlImage(image_path)

    # Caja objetivo: A29:L46 (centrada visualmente según requerimiento)
    box_width, box_height = _range_size_pixels(ws, 'A', 'L', 29, 46)
    if img.width <= 0 or img.height <= 0 or box_width <= 0 or box_height <= 0:
        return

    scale = min(box_width / img.width, box_height / img.height)
    target_width = max(1, int(img.width * scale))
    target_height = max(1, int(img.height * scale))

    x_offset = max(0, (box_width - target_width) // 2)
    y_offset = max(0, (box_height - target_height) // 2)

    img.width = target_width
    img.height = target_height

    marker = AnchorMarker(
        col=column_index_from_string('A') - 1,
        colOff=pixels_to_EMU(x_offset),
        row=29 - 1,
        rowOff=pixels_to_EMU(y_offset),
    )
    size = XDRPositiveSize2D(cx=pixels_to_EMU(target_width), cy=pixels_to_EMU(target_height))
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)


def fill_lista_sheet(ws, items, start_number):
    """Llena una hoja tipo LISTA con los items dados.

    Args:
        ws: Worksheet de openpyxl
        items: lista de dicts con nombre, numero, talla, cantidad, genero
        start_number: número inicial para la columna A (1, 26, 51, ...)
    """
    for i in range(LISTA_ITEMS_PER_SHEET):
        row = LISTA_FIRST_ROW + i
        # Actualizar número en columna A
        ws[f'A{row}'] = start_number + i

        # Corregir tamaño de fuente en B9-B11 (plantilla tiene 14, debe ser 16)
        cell_b = ws[f'B{row}']
        if cell_b.font.size and cell_b.font.size < 16:
            cell_b.font = Font(
                name=cell_b.font.name or 'Aptos Narrow',
                size=16,
                bold=cell_b.font.bold,
                italic=cell_b.font.italic,
                color=cell_b.font.color,
            )

        if i < len(items):
            item = items[i]
            # B = nombre (opcional)
            nombre = item.get('nombre', '')
            if nombre:
                ws[f'B{row}'] = nombre
            # C = número de jugador (opcional)
            numero = item.get('numero', '')
            if numero != '' and numero is not None:
                ws[f'C{row}'] = numero
            # D = talla
            talla = item.get('talla', '')
            if talla:
                ws[f'D{row}'] = talla
            # F = cantidad
            cantidad = item.get('cantidad', '')
            if cantidad != '' and cantidad is not None:
                ws[f'F{row}'] = cantidad
            # G = género
            genero = item.get('genero', '')
            if genero:
                ws[f'G{row}'] = genero
        else:
            # Limpiar celdas de datos si el item no existe
            # (importante para hojas copiadas)
            for col in ['B', 'C', 'D', 'F', 'G']:
                ws[f'{col}{row}'] = None


def generate(input_json_path: str, output_path: str):
    # Leer datos de entrada
    with open(input_json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    template_path = data['template_path']
    folio = data['folio']
    fecha_pedido = data['fecha_pedido']
    cliente = data['cliente']
    cantidad_total = data['cantidad_total']
    tela = data['tela']
    modelo = data['modelo']
    tallas = data.get('tallas', [])
    listado = data.get('listado', [])
    design_image_url = data.get('design_image_url')

    # --- 1. Abrir plantilla con openpyxl ---
    wb = openpyxl.load_workbook(template_path)
    ws = wb['PORTADA']

    # Capturar imágenes de LISTA 1 inmediatamente (antes de que los streams se cierren)
    ws_lista_ref = wb['LISTA 1']
    lista_images_data = []
    for orig_img in ws_lista_ref._images:
        img_bytes = orig_img._data()
        lista_images_data.append({
            'data': img_bytes,
            'anchor': copy.deepcopy(orig_img.anchor),
            'width': orig_img.width,
            'height': orig_img.height,
        })
        # Reemplazar el ref original con un BytesIO fresco para que wb.save() funcione
        orig_img.ref = io.BytesIO(img_bytes)

    # --- 2. Escribir datos principales en PORTADA ---
    ws['L5'] = folio
    ws['L6'] = fecha_pedido
    ws['D9'] = cliente
    ws['K12'] = f'{cantidad_total} PLAYERAS'
    ws['C19'] = tela
    ws['H19'] = modelo

    # --- 3. Escribir tallas en PORTADA ---
    for talla_item in tallas:
        tipo = talla_item['tipo'].upper()
        talla = talla_item['talla'].upper()
        cantidad = talla_item.get('cantidad', 1)

        tipo_map = MAPA_TALLAS.get(tipo)
        if not tipo_map:
            print(f'WARN: Tipo "{tipo}" no válido, saltando', file=sys.stderr)
            continue

        cell_addr = tipo_map.get(talla)
        if not cell_addr:
            print(f'WARN: Talla "{talla}" no válida para tipo "{tipo}", saltando', file=sys.stderr)
            continue

        ws[cell_addr] = cantidad

    # --- 3.1 Insertar imagen del diseno en PORTADA (A29:L46 centrada) ---
    temp_downloaded_image = ''
    if design_image_url:
        try:
            image_path = design_image_url
            if str(design_image_url).startswith('http://') or str(design_image_url).startswith('https://'):
                image_path = _download_image_to_temp(design_image_url)
                temp_downloaded_image = image_path

            if os.path.exists(image_path):
                _insert_centered_image(ws, image_path)
            else:
                print('WARN: No se encontro la imagen de diseno para insertar en PORTADA', file=sys.stderr)
        except Exception as img_error:
            print(f'WARN: No se pudo insertar la imagen de diseno: {img_error}', file=sys.stderr)

    # --- 4. Llenar LISTA 1 (y crear hojas adicionales si necesario) ---
    if listado:
        # Dividir listado en bloques de 25
        chunks = [listado[i:i + LISTA_ITEMS_PER_SHEET]
                   for i in range(0, len(listado), LISTA_ITEMS_PER_SHEET)]

        # Llenar la primera hoja LISTA 1
        ws_lista = wb['LISTA 1']
        ws_lista['E4'] = folio
        ws_lista['E5'] = fecha_pedido
        fill_lista_sheet(ws_lista, chunks[0], start_number=1)

        # Crear hojas adicionales si hay más de 25 items
        for chunk_idx in range(1, len(chunks)):
            sheet_num = chunk_idx + 1
            new_ws = wb.copy_worksheet(ws_lista)
            new_ws.title = f'LISTA {sheet_num}'

            # Copiar imágenes de LISTA 1 a la nueva hoja
            for img_info in lista_images_data:
                new_img = XlImage(io.BytesIO(img_info['data']))
                new_img.anchor = copy.deepcopy(img_info['anchor'])
                new_img.width = img_info['width']
                new_img.height = img_info['height']
                new_ws.add_image(new_img)

            # Escribir folio y fecha
            new_ws['E4'] = folio
            new_ws['E5'] = fecha_pedido

            # Llenar datos con numeración continuada
            start_number = chunk_idx * LISTA_ITEMS_PER_SHEET + 1
            fill_lista_sheet(new_ws, chunks[chunk_idx], start_number)

    # --- 5. Guardar directamente ---
    wb.save(output_path)
    wb.close()

    if temp_downloaded_image and os.path.exists(temp_downloaded_image):
        os.unlink(temp_downloaded_image)

    print(json.dumps({'success': True, 'output': output_path}))


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print(json.dumps({'success': False, 'error': 'Uso: python generate_xlsx.py <input.json> <output.xlsx>'}))
        sys.exit(1)

    try:
        generate(sys.argv[1], sys.argv[2])
    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}))
        sys.exit(1)
